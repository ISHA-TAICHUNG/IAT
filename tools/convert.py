#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
即測即評資料轉檔工具（核心）
- 支援 .mdb (Access) → CSV
- 支援 .xlsx (Excel) → CSV（只保留查詢必要欄位，過濾個資）

用法: python3 convert.py <輸入檔案 1> [<輸入檔案 2> ...]
"""

import sys
import os
import csv
import subprocess
from pathlib import Path

# ====== 資安設定：xlsx 報檢資料（術科）只輸出這 8 個欄位 ======
# 其他個資（地址、電話、email、生日、身障類別等）一律不輸出
#
# 重要更新（2026-04-23）：身分證完整字號不再寫入 CSV。
# IDNO 欄位會被轉成 IDNO_LAST4（只保留末 4 碼），大幅降低個資風險。
# 這樣即使 Drive 資料外洩，攻擊者也拿不到完整身分證。
ALLOWED_XLSX_FIELDS = {
    'IDNO',       # 身分證號（會自動裁切為末 4 碼 → IDNO_LAST4）
    'NAME',       # 姓名
    'AENO',       # 准考證號
    'PNO',        # 職類代碼
    'EGR',        # 級別
    'DSTNG',      # 免試別
    'OPEXDT',     # 術科測試日期
    'OPEXTIME',   # 術科測試時間
}

# 身分證欄位裁切設定：原欄位名 → 新欄位名 + 裁切函式
IDNO_TRANSFORM = {
    'source': 'IDNO',
    'zh_new': '身分證末4碼',
    'en_new': 'IDNO_LAST4',
    'transform': lambda v: str(v).strip()[-4:] if v else '',
}

# ====== 學科報檢資料：只保留公開欄位（電話/MAIL 不輸出）======
# 這份 xlsx 標題在第 3 列，資料從第 4 列開始，沒有英文欄位名
ALLOWED_WRITTEN_FIELDS_ZH = {
    '准考證號', '學科測試編號', '場地', '試場', '場次', '測試日期', '測試時間',
}
# 中文欄位名 → 英文欄位名（GAS 端使用）
WRITTEN_FIELD_MAP = {
    '准考證號': 'aeno',
    '學科測試編號': 'examno',
    '場地': 'venue',
    '試場': 'room',
    '場次': 'distid',
    '測試日期': 'exdate',
    '測試時間': 'extime',
}

# mdb 需要的表格（其他不輸出）
ALLOWED_MDB_TABLES = {'AREADIST', 'AREASEAT'}
# mdb 各表允許欄位（AREAWATCH 含個資不輸出；UpLoadLog 無用）
ALLOWED_MDB_FIELDS = {
    'AREADIST': None,  # None = 全部欄位
    'AREASEAT': None,
}

# PUA 解碼僅對這些「純 ASCII 代碼欄位」執行，避免誤改中文欄位（如 ROOMNAME/AREANAME）
# 中文欄位若含 PUA 字元應保留原狀（不破壞既有中文內容）
PUA_DECODE_COLUMNS = {
    'AENO', 'IDNO', 'PNO', 'EGR', 'DSTNG', 'EXAMKIND', 'YR', 'STP',
    'TESTLOTID', 'SETTESTID', 'AREATYPE', 'OPCD', 'AREANO', 'ROOMNO',
    'DISTID', 'DISTNAME', 'COLUMNID', 'COLUMNNO', 'CAPACITY',
    'WPID', 'EXDATE', 'EXTIME', 'EXTIME2', 'STTIME', 'ZIP', 'TEL',
}


def find_mdbtools():
    """找 mdb-tools 執行檔"""
    for path in ['/opt/homebrew/bin/mdb-export', '/usr/local/bin/mdb-export']:
        if os.path.exists(path):
            return path
    return None


def convert_xlsx(src_path, out_dir):
    """xlsx → csv，只保留安全欄位
    自動辨識兩種報檢資料：
      - 一般報檢資料（術科）：SKT1 sheet，第 1 列中文、第 2 列英文欄位名
      - 學科報檢資料：第 1 列合併標題、第 3 列中文欄位名、第 4 列起資料
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        print(f"❌ 缺少 openpyxl 套件，請安裝：pip3 install openpyxl")
        return False, 0

    src = Path(src_path)

    # 學科報檢資料 → 走獨立的轉換路徑
    if '學科報檢資料' in src.stem:
        return convert_written_xlsx(src_path, out_dir)

    # 輸出檔名：保留「報檢資料」識別，其餘簡化
    if '報檢資料' in src.stem:
        out_name = '報檢資料.csv'
    else:
        out_name = src.stem + '.csv'
    out_path = Path(out_dir) / out_name

    wb = load_workbook(src_path, data_only=True, read_only=True)
    ws = wb['SKT1'] if 'SKT1' in wb.sheetnames else wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        print(f"⚠️  {src.name}: 資料不足")
        return False, 0

    zh_headers = rows[0]      # 中文標題
    en_fields = rows[1]       # 英文欄位名

    # 找出允許欄位的索引
    keep_idx = []
    idno_idx = -1  # 記錄 IDNO 欄位位置（之後要裁切）
    for i, field in enumerate(en_fields):
        if field and str(field).strip().upper() in ALLOWED_XLSX_FIELDS:
            keep_idx.append(i)
            if str(field).strip().upper() == IDNO_TRANSFORM['source']:
                idno_idx = i

    if not keep_idx:
        print(f"⚠️  {src.name}: 找不到預期的欄位")
        return False, 0

    # 準備輸出用的標題（IDNO → 身分證末4碼）
    out_zh_headers = []
    out_en_fields = []
    for i in keep_idx:
        if i == idno_idx:
            out_zh_headers.append(IDNO_TRANSFORM['zh_new'])
            out_en_fields.append(IDNO_TRANSFORM['en_new'])
        else:
            out_zh_headers.append(str(zh_headers[i]) if zh_headers[i] else '')
            out_en_fields.append(str(en_fields[i]) if en_fields[i] else '')

    # 寫入 CSV
    count = 0
    with open(out_path, 'w', newline='', encoding='utf-8-sig') as f:
        w = csv.writer(f)
        # 中文標題行
        w.writerow(out_zh_headers)
        # 英文欄位名行（GAS 要用）
        w.writerow(out_en_fields)
        # 資料行
        for row in rows[2:]:
            if not row or not row[keep_idx[0]]:  # 無資料或無身分證
                continue
            out_row = []
            for i in keep_idx:
                val = row[i] if i < len(row) else None
                if i == idno_idx:
                    # 身分證 → 只留末 4 碼
                    out_row.append(IDNO_TRANSFORM['transform'](val))
                else:
                    out_row.append(str(val) if val is not None else '')
            w.writerow(out_row)
            count += 1

    if idno_idx >= 0:
        print(f"✅ {src.name} → {out_name}（{count} 筆，{len(keep_idx)} 個欄位，🔒 身分證已裁切為末 4 碼）")
    else:
        print(f"✅ {src.name} → {out_name}（{count} 筆，{len(keep_idx)} 個欄位）")
    return True, count


def convert_written_xlsx(src_path, out_dir):
    """學科報檢資料 xlsx → 學科報檢資料.csv
    格式：第 1 列為合併大標題、第 3 列為中文欄位名、第 4 列起為資料。
    只保留公開欄位（AENO、測試日期/時間、場地/試場/場次），電話/MAIL 過濾掉。
    """
    from openpyxl import load_workbook
    src = Path(src_path)
    out_name = '學科報檢資料.csv'
    out_path = Path(out_dir) / out_name

    wb = load_workbook(src_path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 4:
        print(f"⚠️  {src.name}: 資料不足")
        return False, 0

    # 第 3 列（index 2）是中文欄位名
    zh_headers = rows[2]
    keep_idx = []
    keep_en = []
    keep_zh = []
    for i, h in enumerate(zh_headers):
        if h and str(h).strip() in ALLOWED_WRITTEN_FIELDS_ZH:
            keep_idx.append(i)
            keep_zh.append(str(h).strip())
            keep_en.append(WRITTEN_FIELD_MAP[str(h).strip()])

    if not keep_idx:
        print(f"⚠️  {src.name}: 找不到預期的欄位")
        return False, 0

    count = 0
    with open(out_path, 'w', newline='', encoding='utf-8-sig') as f:
        w = csv.writer(f)
        w.writerow(keep_zh)          # 中文標題
        w.writerow(keep_en)           # 英文欄位名（GAS 使用）
        for row in rows[3:]:
            if not row or not row[keep_idx[0]]:  # 無 AENO 的列跳過
                continue
            w.writerow([str(row[i]) if row[i] is not None else '' for i in keep_idx])
            count += 1

    print(f"✅ {src.name} → {out_name}（{count} 筆，{len(keep_idx)} 個欄位，個資已過濾）")
    return True, count


def decode_pua_ascii(text):
    """解碼 mdbtools 讀到的 PUA 字元回原始 ASCII
    規則：
      - 數字/符號：U+E4xx → (U+E4xx - 0xE481) = ASCII
      - 字母：U+E5xx → (U+E5xx - 0xE4F4) = ASCII

    已知 bug 修正（mdbtools Jet 壓縮 Unicode 解碼例外）：
      - U+E4A9 實際為 '8'（naive 解會得到 '('）
      - U+E4AA 實際為 '9'（naive 解會得到 ')'）
    """
    if not text:
        return text
    result = []
    for c in text:
        cp = ord(c)
        # 優先處理已知 bug 映射
        if cp == 0xE4A9:
            result.append('8')
        elif cp == 0xE4AA:
            result.append('9')
        elif 0xE481 <= cp < 0xE500:
            # 數字/符號
            result.append(chr(cp - 0xE481))
        elif 0xE500 <= cp < 0xE600:
            # 字母
            result.append(chr(cp - 0xE4F4))
        else:
            result.append(c)
    return ''.join(result)


def convert_mdb(src_path, out_dir):
    """mdb → 多個 csv，只轉允許的表格，並修正 PUA 字元編碼問題"""
    mdb_export = find_mdbtools()
    if not mdb_export:
        print("❌ 未安裝 mdbtools，請執行：brew install mdbtools")
        return False, 0

    src = Path(src_path)

    # 取得表格清單
    mdb_tables = mdb_export.replace('mdb-export', 'mdb-tables')
    result = subprocess.run([mdb_tables, '-1', src_path], capture_output=True, text=True)
    if result.returncode != 0:
        print(f"❌ 無法讀取 {src.name}")
        return False, 0

    tables = [t.strip() for t in result.stdout.split('\n') if t.strip()]
    count = 0
    for table in tables:
        if table not in ALLOWED_MDB_TABLES:
            continue
        out_path = Path(out_dir) / f"{table}.csv"

        # 先匯出到臨時位置
        tmp_path = str(out_path) + '.tmp'
        with open(tmp_path, 'w', encoding='utf-8') as f:
            r = subprocess.run([mdb_export, src_path, table], stdout=f, stderr=subprocess.PIPE)

        if r.returncode != 0:
            os.unlink(tmp_path)
            continue

        # 只對「允許的純 ASCII 欄位」解碼 PUA，中文欄位保留原狀
        fixed_count = 0
        with open(tmp_path, 'r', encoding='utf-8', newline='') as fin, \
             open(out_path, 'w', encoding='utf-8', newline='') as fout:
            reader = csv.reader(fin)
            writer = csv.writer(fout)
            try:
                headers = next(reader)
            except StopIteration:
                headers = []
            writer.writerow(headers)
            # 決定哪些欄位索引要解碼
            decode_idx = {i for i, h in enumerate(headers)
                          if h.strip().upper() in PUA_DECODE_COLUMNS}
            for row in reader:
                new_row = []
                changed = False
                for i, cell in enumerate(row):
                    if i in decode_idx:
                        decoded = decode_pua_ascii(cell)
                        if decoded != cell:
                            changed = True
                        new_row.append(decoded)
                    else:
                        new_row.append(cell)
                if changed:
                    fixed_count += 1
                writer.writerow(new_row)
        os.unlink(tmp_path)

        with open(out_path, 'r', encoding='utf-8') as f:
            rows = sum(1 for _ in f) - 1
        note = f"（修復 {fixed_count} 筆編碼）" if fixed_count else ""
        print(f"✅ {src.name} → {table}.csv（{rows} 筆）{note}")
        count += 1

    # 列出跳過的表
    skipped = [t for t in tables if t not in ALLOWED_MDB_TABLES]
    if skipped:
        print(f"ℹ️  已跳過（含個資或無用）: {', '.join(skipped)}")

    return count > 0, count


def main():
    if len(sys.argv) < 2:
        print("用法: convert.py <檔案 1> [<檔案 2> ...]")
        sys.exit(1)

    files = sys.argv[1:]
    # 輸出到第一個檔案的目錄
    out_dir = os.path.dirname(os.path.abspath(files[0]))

    success = 0
    total_files = 0
    for src in files:
        total_files += 1
        ext = Path(src).suffix.lower()
        if ext == '.xlsx':
            ok, _ = convert_xlsx(src, out_dir)
            if ok: success += 1
        elif ext == '.mdb':
            ok, _ = convert_mdb(src, out_dir)
            if ok: success += 1
        else:
            print(f"⚠️  跳過（不支援的格式）: {src}")

    print()
    print(f"完成: {success}/{total_files} 個檔案成功")
    print(f"輸出目錄: {out_dir}")
    return 0 if success > 0 else 1


if __name__ == '__main__':
    sys.exit(main())
